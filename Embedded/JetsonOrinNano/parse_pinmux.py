#!/usr/bin/env python3
"""
Parse the NVIDIA Jetson Orin NX/Nano Pinmux Config Template (.xlsm) into CSVs.

Usage:
    python parse_pinmux.py [path/to/file.xlsm] [--outdir output/]

Outputs one CSV per sheet plus a combined summary of the chip-down template.
"""

import argparse
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
import pandas as pd

XLSM_DEFAULT = Path(__file__).parent.parent.parent.parent.parent / (
    "Data/Public/embedded/NVIDIAJetsonOrinNano/"
    "Jetson_Orin_NX_and_Orin_Nano_series_Pinmux_Config_Template_v1.2.xlsm"
)

# Sheets worth exporting verbatim (data-heavy, not just prose)
DATA_SHEETS = {
    "E2421_E2425_SLT_Pinmux_Config",
    "Jetson Orin Nano&NX Pinmux DP",
    "Jetson Orin Nano&NX Pinmux HDMI",
    "T234_Chip_Down_Pinmux_Template",
}


def sheet_to_df(ws) -> pd.DataFrame:
    """Read an openpyxl worksheet into a DataFrame, skipping fully-blank rows."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    # Find first non-blank row to use as header
    header_idx = 0
    for i, row in enumerate(rows):
        if any(c is not None for c in row):
            header_idx = i
            break

    header = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(rows[header_idx])]
    data_rows = []
    for row in rows[header_idx + 1 :]:
        if any(c is not None for c in row):
            data_rows.append(row)

    df = pd.DataFrame(data_rows, columns=header)
    # Drop columns that are entirely None
    df = df.dropna(axis=1, how="all")
    return df


def find_chip_down_header(ws) -> int:
    """
    T234_Chip_Down_Pinmux_Template has several title rows before the actual
    pin table. Locate the row where 'Ball Name' or 'Pin Group' appears.
    """
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [str(c).strip() if c is not None else "" for c in row]
        joined = " ".join(cells).lower()
        if "ball name" in joined or "pin group" in joined or "signal name" in joined:
            return i
    return 0


def parse_chip_down(ws, outdir: Path) -> None:
    """Special handling for the main chip-down template sheet."""
    header_row = find_chip_down_header(ws)
    rows = list(ws.iter_rows(values_only=True))
    if header_row >= len(rows):
        print("  WARNING: could not find header row in chip-down sheet")
        return

    # Build multi-level header: NVIDIA uses 2-3 merged header rows sometimes;
    # just flatten them by joining non-None cells across the first 3 rows.
    header_rows = rows[header_row : header_row + 2]
    merged = []
    for col_idx in range(len(rows[header_row])):
        parts = []
        for hrow in header_rows:
            val = hrow[col_idx] if col_idx < len(hrow) else None
            if val is not None:
                s = str(val).strip()
                if s and s not in parts:
                    parts.append(s)
        merged.append(" / ".join(parts) if parts else f"col_{col_idx}")

    data_start = header_row + len(header_rows)
    data_rows = []
    for row in rows[data_start:]:
        if any(c is not None for c in row):
            data_rows.append(row)

    df = pd.DataFrame(data_rows, columns=merged)
    df = df.dropna(axis=1, how="all")

    out = outdir / "T234_chip_down_pinmux.csv"
    df.to_csv(out, index=False)
    print(f"  → {out}  ({len(df)} pins, {len(df.columns)} columns)")

    # Convenience: emit a compact summary (Ball Name + mux options only)
    ball_col = next((c for c in df.columns if "ball" in c.lower()), None)
    if ball_col:
        mux_cols = [c for c in df.columns if "sfio" in c.lower() or "alt" in c.lower() or "function" in c.lower()]
        summary_cols = [c for c in ([ball_col] + mux_cols) if c in df.columns]
        if summary_cols:
            out_s = outdir / "T234_chip_down_pinmux_summary.csv"
            df[summary_cols].to_csv(out_s, index=False)
            print(f"  → {out_s}  (ball + mux functions only)")


def extract_ball_names_xml(xlsm_path: Path, outdir: Path) -> None:
    """
    Extract Verilog Ball Names directly from the SLT sheet XML, bypassing the
    #REF! formula cells. Column A (Verilog Ball Name) is always a plain string,
    so this gives us the complete authoritative list of all T234 pads.
    """
    NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(xlsm_path) as z:
        sst = ET.fromstring(z.read("xl/sharedStrings.xml"))
        strings = [
            "".join(t.text or "" for t in si.findall(".//x:t", NS))
            for si in sst.findall("x:si", NS)
        ]
        # SLT sheet is sheet1.xml per workbook rels
        ws = ET.fromstring(z.read("xl/worksheets/sheet1.xml"))

    HEADER_ROW = 7  # 1-indexed row where 'Verilog Ball Name' header lives
    DATA_START = 9   # first actual pin row
    ball_names = []
    for row_el in ws.findall(".//x:row", NS):
        rnum = int(row_el.get("r", 0))
        if rnum < DATA_START:
            continue
        # Column A cell is 'A<rnum>'
        for c_el in row_el.findall("x:c", NS):
            if c_el.get("r", "").startswith("A"):
                t = c_el.get("t", "n")
                v_el = c_el.find("x:v", NS)
                if v_el is not None and v_el.text:
                    val = strings[int(v_el.text)] if t == "s" else v_el.text
                    if val and val != "#REF!":
                        ball_names.append({"row": rnum, "verilog_ball_name": val})
                break

    df = pd.DataFrame(ball_names)
    out = outdir / "t234_ball_names.csv"
    df.to_csv(out, index=False)
    print(f"  → {out}  ({len(df)} ball/pad names)")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsm", nargs="?", default=str(XLSM_DEFAULT), help="Path to .xlsm file")
    parser.add_argument("--outdir", default="output", help="Output directory for CSVs")
    args = parser.parse_args()

    xlsm_path = Path(args.xlsm)
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    if not xlsm_path.exists():
        sys.exit(f"ERROR: file not found: {xlsm_path}")

    print(f"Loading {xlsm_path.name} …")
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, keep_vba=False, data_only=True)
    print(f"Sheets: {wb.sheetnames}\n")

    for name in wb.sheetnames:
        ws = wb[name]
        safe = name.replace("/", "_").replace("&", "and").replace(" ", "_")
        print(f"[{name}]")

        if name == "T234_Chip_Down_Pinmux_Template":
            parse_chip_down(ws, outdir)
        elif name in DATA_SHEETS:
            df = sheet_to_df(ws)
            if df.empty:
                print("  (empty)")
            else:
                out = outdir / f"{safe}.csv"
                df.to_csv(out, index=False)
                print(f"  → {out}  ({len(df)} rows, {len(df.columns)} cols)")
        else:
            # Prose/readme sheets — still dump but note them
            df = sheet_to_df(ws)
            if not df.empty:
                out = outdir / f"{safe}.csv"
                df.to_csv(out, index=False)
                print(f"  → {out}  (prose/guide, {len(df)} rows)")
            else:
                print("  (empty)")
        print()

    wb.close()

    # Extract ball names directly from XML (formula cells are broken, col A is plain text)
    print("[Ball names via raw XML]")
    extract_ball_names_xml(xlsm_path, outdir)
    print()

    print("Done. CSVs written to:", outdir.resolve())


if __name__ == "__main__":
    main()
